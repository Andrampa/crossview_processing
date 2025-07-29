import pandas as pd
import json, os, re, glob, requests
from arcgis.gis import GIS
from arcgis.features import FeatureLayer
import pandas as pd
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage


# Path for storing and reusing the survey_list
survey_list_cache_path = os.path.join(os.path.dirname(__file__), "survey_list_cache.json")

######### get updated list of surveys to process ###########
# Filter to keep only selected countries
selected_countries = [
    "AFG",  # Afghanistan
    "BGD",  # Bangladesh
    "BFA",  # Burkina Faso
    "CMR",  # Cameroon
    "CAF",  # Central African Republic (CAR)
    "TCD",  # Chad
    "COL",  # Colombia
    "COD",  # Democratic Republic of the Congo (DRC)
    "SLV",  # El Salvador
    "GTM",  # Guatemala
    "HTI",  # Haiti
    "HND",  # Honduras
    "LBN",  # Lebanon
    "MLI",  # Mali
    "MOZ",  # Mozambique
    "MMR",  # Myanmar
    "NER",  # Niger
    "NGA",  # Nigeria
    "PSE",  # West Bank (Palestine)
    "YEM",  # Yemen
    "IRQ",  # Iraq
    "PAK",  # Pakistan
    "MWI",  # Malawi
    "ZWE"   # Zimbabwe
]

# selected_countries = [
#    'YEM']

get_updated_list_of_surveys_from_AGOL = True
dev_mode = False

# Create the output directory if it does not exist
output_dir = "outputs_for_erps"
os.makedirs(output_dir, exist_ok=True)

if not get_updated_list_of_surveys_from_AGOL:
    print("Getting updated_list_of_surveys_from cache")
    # Load cached survey list from JSON file
    if os.path.exists(survey_list_cache_path):
        with open(survey_list_cache_path, "r", encoding="utf-8") as f:
            survey_list = json.load(f)
    else:
        raise FileNotFoundError("Cached survey list not found. Please set get_updated_list_of_surveys_from_AGOL = True once to fetch and store it.")

else:
    print("Getting updated_list_of_surveys_from_agol")
    # Feature layer URL (Layer 0)
    layer_url = "https://services5.arcgis.com/sjP4Ugu5s0dZWLjd/arcgis/rest/services/OER_Monitoring_System_View/FeatureServer/0"
    layer = FeatureLayer(layer_url)
    # Query all features
    features = layer.query(where="round_validated='Yes'", out_fields="admin0_isocode, round, admin0_name_en, coll_end_date, card1_indicator2, card1_indicator2_text, methodology",   return_geometry=False)

    # Convert to DataFrame
    df = features.sdf
    # Drop duplicates to get unique combinations
    unique_combinations = df.drop_duplicates(subset=["admin0_isocode", "round"]).reset_index(drop=True)
    # Clean and prepare round column
    df_clean = df.dropna(subset=["admin0_isocode", "round"]).copy()
    # Extract numeric part from "Round 01", "Round 12", etc.
    df_clean["round_num"] = df_clean["round"].str.extract(r"Round\s+(\d+)", expand=False).astype(int)
    # Get the highest round per country
    latest_rounds = df_clean.sort_values("round_num", ascending=False).drop_duplicates(subset=["admin0_isocode"])
    # Sort by country code
    latest_rounds = latest_rounds.sort_values("admin0_isocode").reset_index(drop=True)

    latest_rounds = latest_rounds[latest_rounds["admin0_isocode"].isin(selected_countries)].reset_index(drop=True)
    latest_rounds = latest_rounds.rename(columns={"admin0_isocode": "adm0_iso3", "admin0_name_en": "adm0_name", "card1_indicator2": "diem_survey_coverage", "card1_indicator2_text": "diem_target_pop"})
    survey_list = latest_rounds.to_dict(orient="records")
    # Print the result
    # for survey in survey_list:
    #     print(survey["adm0_iso3"], survey["round_num"], survey.get("coll_end_date", "N/A"))
    # Save survey list to cache file
    with open(survey_list_cache_path, "w", encoding="utf-8") as f:
        json.dump(survey_list, f, indent=2, default=str)


    # Export to Excel file THE RECAP ON COVERAGE AND METHODOLOGY
    excel_columns = [
        "adm0_iso3",
        "adm0_name",
        "round",
        "coll_end_date",
        "diem_target_pop",
        "diem_survey_coverage",
        "methodology"
    ]

    latest_rounds["coll_end_date"] = pd.to_datetime(latest_rounds["coll_end_date"], errors="coerce").dt.strftime(
        "%b %Y")

    # Ensure all selected columns are present and ordered
    export_df = latest_rounds.rename(columns={
        "round": "round",  # Keep name consistent
        "coll_end_date": "coll_end_date",
        "methodology": "methodology"
    })[excel_columns]

    # Define path inside the subdirectory
    output_excel_path = os.path.join(output_dir, f"DIEM_latest_surveys_coverage_and_methodology.xlsx")

    # Save to Excel
    export_df.to_excel(output_excel_path, index=False)
    print(f"Excel file exported: {output_excel_path}")


# Show all columns and full width when printing DataFrames
pd.set_option("display.max_columns", None)
pd.set_option("display.width", None)
pd.set_option("display.max_colwidth", None)

# === PARAMETERS ===

group_indicators = {
    "fies_resid": ["hh_residencetype", "Residence type" ],
    "fies_hhtype": ["hh_agricactivity", "Agricultural activity" ],
    "agriculture": ["hh_residencetype", "Residence type" ]}
shock_label_map = {
    "shock_coldtemporhail": "Cold temperature or hail",
    "shock_sicknessordeathofhh": "Sickness or death in household",
    "shock_napasture": "No access to pasture",
    "shock_higherfoodprices": "Higher food prices",
    "shock_lostemplorwork": "Loss of employment or income",
    "shock_landslides": "Landslides",
    "shock_animaldisease": "Animal disease",
    "shock_drought": "Drought",
    "shock_plantdisease": "Plant disease",
    "shock_flood": "Flood",
    "shock_higherfuelprices": "Higher fuel prices",
    "shock_earthquake": "Earthquake",
    "shock_firemanmade": "Manmade fire",
    "shock_mvtrestrict": "Movement restrictions",
    "shock_pestoutbreak": "Pest outbreak",
    "shock_theftofprodassets": "Theft of productive assets",
    "shock_firenatural": "Natural fire",
    "shock_violenceinsecconf": "Violence, insecurity or conflict"
}

export_csv = True

# === LOAD INDICATOR METADATA ===
with open("indicators_metadata.js", "r", encoding="utf-8") as f:
    js_text = f.read()
js_clean = re.sub(r"^window\.indicatorData\s*=\s*", "", js_text.strip())
if js_clean.endswith(";"):
    js_clean = js_clean[:-1]
indicator_data = json.loads(js_clean)

def get_indicator_info(indicator_name):
    if indicator_name not in indicator_data:
        raise ValueError(f"Indicator '{indicator_name}' not found in metadata")
    return {
        "title": indicator_data[indicator_name].get("title", indicator_name),
        "codes": indicator_data[indicator_name]["codes"]
    }


# === DEFINE FUNCTION  ===

def fies_by_indicator(indicator_key, df, universe_label=None):
    fies_fields = {
        "p_mod": "moderate/severe (p_mod)",
        "p_sev": "severe only (p_sev)"
    }
    fies_rows = []

    indicator, custom_label = group_indicators[indicator_key]  # NEW: extract both the field name and axis label
    info = get_indicator_info(indicator)
    title = custom_label  # use the custom label as table index and x-axis label
    codes = info["codes"]

    # Generate the descriptive subtitle
    if universe_label is None:
        universe_label = "all households"
    subtitle = f"FIES weighted average by {custom_label} — Universe: {universe_label}"

    for code, label in codes.items():
        if label in ["Don't know", "Refused"]:
            continue
        try:
            group_df = df[df[indicator].astype("Int64") == int(code)]
        except ValueError:
            group_df = df[df[indicator].astype(str) == str(code)]

        if len(group_df) == 0:
            continue

        row_data = {}
        for field, label_field in fies_fields.items():
            weighted_sum = 0.0
            total_weight = 0.0
            for _, row in group_df.iterrows():
                try:
                    w = float(row.get("weight_final", 0))
                    val = float(row.get(field, 0))
                    if w > 0 and pd.notna(val):
                        weighted_sum += w * val
                        total_weight += w
                except:
                    continue
            row_data[label_field] = round((weighted_sum / total_weight) * 100, 1) if total_weight > 0 else 0.0
        row_data[custom_label] = label
        fies_rows.append(row_data)

    fies_df = pd.DataFrame(fies_rows).set_index(custom_label)

    # TOTAL row
    total_row = {}
    for field, label_field in fies_fields.items():
        weighted_sum = 0.0
        total_weight = 0.0
        for _, row in df.iterrows():
            try:
                w = float(row.get("weight_final", 0))
                val = float(row.get(field, 0))
                if w > 0 and pd.notna(val):
                    weighted_sum += w * val
                    total_weight += w
            except:
                continue
        total_row[label_field] = round((weighted_sum / total_weight) * 100, 1) if total_weight > 0 else 0.0
    fies_df.loc["All"] = pd.Series(total_row)

    # Custom title for table/chart (depends only on indicator_key)
    if indicator_key == 'fies_hhtype':
        title = "FIES: Prevalence of recent food insecurity by agricultural dependency (detailed)"
    elif indicator_key == 'fies_resid':
        title = "FIES: Prevalence of recent food insecurity by residency status"
    else:
        title = f"FIES: Prevalence of recent food insecurity by {custom_label}"

    return {
        "title": title,
        "subtitle": subtitle,
        "df": fies_df
    }

def fies_by_simplified_agriculture(df):
    df = df.copy()
    df.loc[:, "agriculture_group"] = df["hh_agricactivity"].apply(lambda x: (
        "Agricultural HH" if int(x) in [1, 2, 3] else
        "Non-agricultural HH" if int(x) == 4 else
        "Don't know" if int(x) == 888 else
        "Refused" if int(x) == 999 else None
    ) if pd.notna(x) else None)

    fies_fields = {
        "p_mod": "moderate/severe (p_mod)",
        "p_sev": "severe only (p_sev)"
    }
    rows = []
    for group in df["agriculture_group"].dropna().unique():
        if group in ["Don't know", "Refused"]:
            continue  # Exclude these categories from the output table
        group_df = df[df["agriculture_group"] == group]
        if len(group_df) == 0:
            continue
        row_data = {}
        for field, label_field in fies_fields.items():
            weighted_sum = 0.0
            total_weight = 0.0
            for _, row in group_df.iterrows():
                try:
                    w = float(row.get("weight_final", 0))
                    val = float(row.get(field, 0))
                    if w > 0 and pd.notna(val):
                        weighted_sum += w * val
                        total_weight += w
                except:
                    continue
            row_data[label_field] = round((weighted_sum / total_weight) * 100, 1) if total_weight > 0 else 0.0
        row_data["Agriculture group"] = group
        rows.append(row_data)

    fies_df = pd.DataFrame(rows).set_index("Agriculture group")

    total_row = {}
    for field, label_field in fies_fields.items():
        weighted_sum = 0.0
        total_weight = 0.0
        for _, row in df.iterrows():
            try:
                w = float(row.get("weight_final", 0))
                val = float(row.get(field, 0))
                if w > 0 and pd.notna(val):
                    weighted_sum += w * val
                    total_weight += w
            except:
                continue
        total_row[label_field] = round((weighted_sum / total_weight) * 100, 1) if total_weight > 0 else 0.0
    fies_df.loc["All"] = pd.Series(total_row)
    return {
        "title": "FIES: Prevalence of recent food insecurity by agricultural dependency",
        "df": fies_df
    }


def agricultural_dependency(df):
    # Full category labels
    agriculture_categories = {
        1: "Yes - crop production",
        2: "Yes - livestock production",
        3: "Yes - both crop and livestock production",
        4: "Non-agricultural household",
        888: "Don't know",
        999: "Refused"
    }

    # Keep only valid analytical codes (exclude DK and REF)
    agriculture_keys = [1, 2, 3, 4]


    agri_indicator, agri_title = group_indicators["agriculture"]
    agri_info = get_indicator_info(agri_indicator)
    agri_codes = agri_info["codes"]
    rows = []

    for code, label in agri_codes.items():
        if label in ["Don't know", "Refused"]:
            continue
        try:
            group_df = df[df[agri_indicator].astype("Int64") == int(code)]
        except ValueError:
            group_df = df[df[agri_indicator].astype(str) == str(code)]

        if len(group_df) == 0:
            continue
        #print(f"Processing AGRICULTURE for group '{label}' ({len(group_df)} records)...")

        weighted_counts = {k: 0.0 for k in agriculture_keys}
        total_weight = 0.0
        for _, row in group_df.iterrows():
            try:
                w = float(row.get("weight_final"))
                val = int(row.get("hh_agricactivity"))
                if w > 0 and val in agriculture_keys:
                    weighted_counts[val] += w
                    total_weight += w
            except:
                continue

        row_data = {
            agriculture_categories[k]: (weighted_counts[k] / total_weight * 100 if total_weight > 0 else 0.0)
            for k in agriculture_keys
        }
        row_data[agri_title] = label
        rows.append(row_data)

    df_result = pd.DataFrame(rows).set_index(agri_title).round(1)

    # TOTAL
    total_counts = {k: 0.0 for k in agriculture_keys}
    total_weight = 0.0
    for _, row in df.iterrows():
        try:
            w = float(row.get("weight_final"))
            val = int(row.get("hh_agricactivity"))
            if w > 0 and val in agriculture_keys:
                total_counts[val] += w
                total_weight += w
        except:
            continue

    total_row = {
        agriculture_categories[k]: (total_counts[k] / total_weight * 100 if total_weight > 0 else 0.0)
        for k in agriculture_keys
    }
    df_result.loc["All"] = pd.Series(total_row).round(1)

    return {
        "title": "Detailed agricultural dependency by residency status",
        "df": df_result
    }

def simplified_dependency_by_residency(df):
    df = df.copy()
    df.loc[:, "agri_simple"] = df["hh_agricactivity"].apply(lambda x: (

        "Agricultural HH" if int(x) in [1, 2, 3] else
        "Non-agricultural HH" if int(x) == 4 else None
    ) if pd.notna(x) else None)

    agri_indicator, resid_title = group_indicators["fies_resid"]
    resid_info = get_indicator_info(agri_indicator)
    resid_codes = resid_info["codes"]
    rows = []

    for code, label in resid_codes.items():
        if label in ["Don't know", "Refused"]:
            continue
        try:
            group_df = df[df[agri_indicator].astype("Int64") == int(code)]
        except ValueError:
            group_df = df[df[agri_indicator].astype(str) == str(code)]
        #group_df = df[df[agri_indicator].astype(str) == code]
        if len(group_df) == 0:
            continue
        #print(f"Processing SIMPLIFIED AGRIC DEPENDENCY for residency group '{label}' ({len(group_df)} records)...")
        weighted_counts = {"Agricultural HH": 0.0, "Non-agricultural HH": 0.0}
        total_weight = 0.0
        for _, row in group_df.iterrows():
            group = row.get("agri_simple")
            try:
                w = float(row.get("weight_final", 0))
                if w > 0 and group in weighted_counts:
                    weighted_counts[group] += w
                    total_weight += w
            except:
                continue
        row_data = {
            "Agricultural HH": (weighted_counts["Agricultural HH"] / total_weight * 100 if total_weight > 0 else 0.0),
            "Non-agricultural HH": (weighted_counts["Non-agricultural HH"] / total_weight * 100 if total_weight > 0 else 0.0),
            resid_title: label
        }
        rows.append(row_data)

    df_result = pd.DataFrame(rows).set_index(resid_title).round(1)

    # TOTAL
    total_counts = {"Agricultural HH": 0.0, "Non-agricultural HH": 0.0}
    total_weight = 0.0
    for _, row in df.iterrows():
        group = row.get("agri_simple")
        try:
            w = float(row.get("weight_final", 0))
            if w > 0 and group in total_counts:
                total_counts[group] += w
                total_weight += w
        except:
            continue
    total_row = {
        "Agricultural HH": (total_counts["Agricultural HH"] / total_weight * 100 if total_weight > 0 else 0.0),
        "Non-agricultural HH": (total_counts["Non-agricultural HH"] / total_weight * 100 if total_weight > 0 else 0.0)
    }
    df_result.loc["All"] = pd.Series(total_row).round(1)

    return {
        "title": "Simplified agricultural dependency by residency status",
        "df": df_result
    }

def needs_summary_grouped(df_all, adm0_iso3, round_num, use_grouping=True, use_previous_round=True, universe_filter=[1]):
    import collections

    # Choose round (current or previous)
    selected_round = round_num - 1 if use_previous_round else round_num

    # Filter to selected round, country, and universe
    df_universe = df_all[
        (df_all["adm0_iso3"] == adm0_iso3) &
        (df_all["round"] == selected_round) &
        (df_all["need"].isin(universe_filter))
    ].copy()

    universe_count = len(df_universe)

    # Total weight
    total_weight = df_universe["weight_final"].apply(pd.to_numeric, errors="coerce").dropna()
    total_weight = total_weight[total_weight > 0].sum()

    # Reclassification
    need_fields_dict = {
        "need_food": "need_food",
        "need_cash": "need_cash",
        "need_other": "agricultural livelihoods",
        "need_vouchers_fair": "agricultural livelihoods",
        "need_crop_inputs": "agricultural livelihoods",
        "need_crop_infrastructure": "agricultural livelihoods",
        "need_crop_knowledge": "agricultural livelihoods",
        "need_ls_feed": "agricultural livelihoods",
        "need_ls_vet_service": "agricultural livelihoods",
        "need_ls_infrastructure": "agricultural livelihoods",
        "need_ls_knowledge": "agricultural livelihoods",
        "need_fish_inputs": "agricultural livelihoods",
        "need_fish_infrastructure": "agricultural livelihoods",
        "need_fish_knowledge": "agricultural livelihoods",
        "need_env_infra_rehab": "agricultural livelihoods",
        "need_cold_storage": "agricultural livelihoods",
        "need_marketing_supp": "agricultural livelihoods"
    }

    # Grouping logic
    if use_grouping:
        from collections import defaultdict
        group_fields = defaultdict(list)
        for field, group in need_fields_dict.items():
            group_fields[group].append(field)
    else:
        all_fields = list(need_fields_dict.keys())
        group_fields = {field: [field] for field in all_fields}

    # Initialize group sums
    group_weight_sums = {group: 0.0 for group in group_fields}

    for _, row in df_universe.iterrows():
        try:
            w = float(row.get("weight_final", 0))
            if not w > 0:
                continue
        except:
            continue

        for group, fields in group_fields.items():
            for field in fields:
                val = row.get(field)
                if val in [1, "1", True]:
                    group_weight_sums[group] += w
                    break  # Only once per group

    # Format output
    results = []
    for group, w_sum in group_weight_sums.items():
        pct = (w_sum / total_weight * 100) if total_weight > 0 else 0.0
        label = group.replace("need_", "").replace("_", " ").capitalize()
        results.append({
            "Need group": label,
            "Weighted % of HH": round(pct, 1)
        })

    df_result = pd.DataFrame(results).set_index("Need group")

    # Dynamic title
    title = (
        f"Needs reported in previous round\n"
        f"Round: {selected_round}, Country: {adm0_iso3.upper()}, Universe: need in {universe_filter}"
    )

    return {
        "title": "Needs reported in previous round",
        "metadata": f"Round: {selected_round}, Country: {adm0_iso3.upper()}, Universe: need in {universe_filter}",
        "df": df_result
    }

def assistance_summary(df, use_grouping=True, universe_filter=[1], round_num=None, adm0_iso3=None):
    import collections

    df_universe = df[df["need"].isin(universe_filter)].copy()
    universe_count = len(df_universe)
    total_weight = df_universe["weight_final"].apply(pd.to_numeric, errors="coerce").dropna()
    total_weight = total_weight[total_weight > 0].sum()

    assistance_provided_dict = {
        "need_received_food": "need_received_food",
        "need_received_cash": "need_received_cash",
        "need_received_vouchers_fair": "agricultural livelihoods",
        "need_received_crop_assist": "agricultural livelihoods",
        "need_received_ls_assist": "agricultural livelihoods",
        "need_received_fish_assist": "agricultural livelihoods",
        "need_received_rehabilitation": "agricultural livelihoods",
        "need_received_sales_support": "agricultural livelihoods",
        "need_received_other": "agricultural livelihoods",
        "need_received_none": "need_received_none"
    }

    if use_grouping:
        group_fields = collections.defaultdict(list)
        for field, group in assistance_provided_dict.items():
            group_fields[group].append(field)
    else:
        all_fields = list(assistance_provided_dict.keys())
        group_fields = {field: [field] for field in all_fields}

    group_weight_sums = {group: 0.0 for group in group_fields.keys()}
    for _, row in df_universe.iterrows():
        try:
            w = float(row.get("weight_final", 0))
            if not w > 0:
                continue
        except:
            continue
        for group, fields in group_fields.items():
            for field in fields:
                val = row.get(field)
                if val in [1, "1", True]:
                    group_weight_sums[group] += w
                    break

    results = []
    for group, w_sum in group_weight_sums.items():
        pct = (w_sum / total_weight * 100) if total_weight > 0 else 0.0
        if group == "need_received_food":
            label = "Food"
        elif group == "need_received_cash":
            label = "Cash"
        else:
            label = group.replace("need_received_", "").replace("_", " ").capitalize()

        results.append({
            "Assistance type": label,
            "Weighted % of HH": round(pct, 1)
        })

    df_result = pd.DataFrame(results).set_index("Assistance type")

    # Compose detailed title
    title = (
        f"Assistance received in round {round_num}\n"
        f"Country: {adm0_iso3.upper() if adm0_iso3 else 'N/A'}, Universe: need in {universe_filter}"
    )

    return {
        "title": f"Assistance received in round {round_num}",
        "metadata": f"Country: {adm0_iso3.upper() if adm0_iso3 else 'N/A'}, Universe: need in {universe_filter}",
        "df": df_result
    }

def compare_needs_vs_assistance(needs_result, assistance_result):
    # Extract and align dataframes
    df_needs = needs_result["df"].copy()
    df_assistance = assistance_result["df"].copy()

    # Rename value columns and drop universe columns
    df_needs = df_needs.rename(columns={
        "Weighted % of HH": "Need in previous round (%)"
    }).drop(columns=["Universe (n)"], errors="ignore")

    df_assistance = df_assistance.rename(columns={
        "Weighted % of HH": "Assistance received (%)"
    }).drop(columns=["Universe (n)"], errors="ignore")

    # Align indices
    df_needs.index.name = "Type"
    df_assistance.index.name = "Type"

    # Join both tables
    merged_df = df_needs.join(df_assistance, how="outer").fillna(0).round(1)

    return {
        "title": "Comparison: Needs in previous round vs Assistance received",
        "metadata": f"Needs → Round {round_num - 1}, Assistance → Round {round_num}, Country: {adm0_iso3.upper()}, Universe: need in [0, 1, 888]",
        "df": merged_df
    }

def assistance_quality_summary(df, group_by="need_received"):
    from collections import defaultdict

    # === Satisfaction categories (full, but 888 and 999 will be excluded later) ===
    assistance_categories = {
        1: "Yes - satisfied",
        2: "No - not received on time",
        3: "No - did not meet my needs",
        4: "No - quantity was not sufficient",
        5: "No - problem with provider",
        6: "No - malfunction of in-kind assistance",
        888: "Don't know",
        999: "Refused"
    }

    # === Grouping logic ===
    if group_by == "need_received":
        field_dict = {
            "need_received_food": "need_received_food",
            "need_received_cash": "need_received_cash",
            "need_received_vouchers_fair": "agricultural livelihoods",
            "need_received_crop_assist": "agricultural livelihoods",
            "need_received_ls_assist": "agricultural livelihoods",
            "need_received_fish_assist": "agricultural livelihoods",
            "need_received_rehabilitation": "agricultural livelihoods",
            "need_received_sales_support": "agricultural livelihoods",
            "need_received_other": "need_received_other",
            "need_received_none": "need_received_none",
            "need_received_dk": "need_received_dk",
            "need_received_ref": "need_received_ref"
        }
        group_label = "Received"
    elif group_by == "need":
        field_dict = {
            "need_food": "need_food",
            "need_cash": "need_cash",
            "need_other": "agricultural livelihoods",
            "need_dk": "need_dk",
            "need_ref": "need_ref",
            "need_vouchers_fair": "agricultural livelihoods",
            "need_crop_inputs": "agricultural livelihoods",
            "need_crop_infrastructure": "agricultural livelihoods",
            "need_crop_knowledge": "agricultural livelihoods",
            "need_ls_feed": "agricultural livelihoods",
            "need_ls_vet_service": "agricultural livelihoods",
            "need_ls_infrastructure": "agricultural livelihoods",
            "need_ls_knowledge": "agricultural livelihoods",
            "need_fish_inputs": "agricultural livelihoods",
            "need_fish_infrastructure": "agricultural livelihoods",
            "need_fish_knowledge": "agricultural livelihoods",
            "need_env_infra_rehab": "agricultural livelihoods",
            "need_cold_storage": "agricultural livelihoods",
            "need_marketing_supp": "agricultural livelihoods"
        }
        group_label = "Need"
    else:
        raise ValueError("group_by must be either 'need_received' or 'need'")

    group_fields = defaultdict(list)
    for field, group in field_dict.items():
        group_fields[group].append(field)

    # Only use the valid categories for output
    valid_categories = [1, 2, 3, 4, 5, 6]

    result_rows = []

    for group, fields in group_fields.items():
        group_df = df[df[fields].isin([1, "1", True]).any(axis=1)].copy()
        if group_df.empty:
            continue

        #print(f"Processing ASSISTANCE QUALITY for group '{group}' ({len(group_df)} records)...")

        weighted_counts = {cat: 0.0 for cat in valid_categories}
        total_weight = 0.0

        for _, row in group_df.iterrows():
            try:
                w = float(row.get("weight_final", 0))
                val = int(row.get("assistance_quality"))
                if w > 0 and val in weighted_counts:
                    weighted_counts[val] += w
                    total_weight += w
            except:
                continue

        if total_weight == 0:
            continue

        row = {
            f"{group_label} group": group.replace("need_received_", "").replace("need_", "").replace("_", " ").capitalize()
        }
        for cat in valid_categories:
            label = assistance_categories[cat]
            row[label] = round((weighted_counts[cat] / total_weight) * 100, 1)

        result_rows.append(row)

    index_name = f"{group_label} group"
    df_result = pd.DataFrame(result_rows).set_index(index_name).round(1)

    sample_size = df["assistance_quality"].dropna().shape[0]
    warning = " — Warning: small sample size" if sample_size < 100 else ""
    title = f"Satisfaction with assistance by group of assistance received (sample for question on satisfaction = {sample_size}){warning}"

    return {
        "title": title,
        "df": df_result
    }


def extract_top10_by_cropland(country_iso3, file_name="IPC_multicountry_20250707.xlsx"):
    """
    Loads the IPC Excel file, extracts the sheet for the given country (ISO3),
    and returns a dictionary with a title and the top 10 adm2 units
    with the highest cropland exposed to floods (rounded to 1 decimal).
    Unnecessary fields are removed and adm2 fields are listed before adm1.
    If the sheet is missing, empty, or an error occurs, 'SKIPPED' is included in the title.
    """
    base_title = "List of adm2 units in IPC3+ with the highest amount of cropland exposed to floods"

    columns_to_keep = [
        "adm0_name", "adm0_ISO3", "adm1_name", "adm1_pcode", "adm2_name", "adm2_pcode",
        "cropland_exp_sqkm", "cropland_tot_sqkm", "crop_exp_perc",
        "pop_exposed", "pop_total", "pop_exp_perc",
        "area_phase_current", "analysis_period_current",
        "area_phase_proj1", "analysis_period_proj1",
        "area_phase_proj2", "analysis_period_proj2"  # <-- ADD THESE
    ]

    rename_dict = {
        "cropland_exp_sqkm": "Cropland exposed (Km2)",
        "cropland_tot_sqkm": "Total cropland (Km2)",
        "crop_exp_perc": "% of cropland exposed",
        "pop_exposed": "Population exposed",
        "pop_total": "Total population",
        "pop_exp_perc": "% of population exposed",
        "area_phase_current": "IPC/CH area current",
        "analysis_period_current": "Period of current IPC analysis",
        "area_phase_proj1": "IPC/CH area projected",
        "analysis_period_proj1": "Period of projected IPC analysis",
        "area_phase_proj2": "IPC/CH area projected 2nd",  # <-- ADD
        "analysis_period_proj2": "Period of 2nd projected IPC analysis"  # <-- ADD
    }

    try:
        file_path = os.path.join(os.path.dirname(__file__), file_name)
        df = pd.read_excel(file_path, sheet_name=country_iso3)

        if df.empty:
            print(f"Sheet '{country_iso3}' exists but is empty.")
            return {
                "title": f"{base_title} — SKIPPED: Sheet '{country_iso3}' is empty",
                "df": pd.DataFrame()
            }

        # Keep only the columns that exist in the sheet
        available_columns = [col for col in columns_to_keep if col in df.columns]
        df_filtered = df[available_columns].rename(columns=rename_dict)

        # Drop area_phase_proj2 and analysis_period_proj2 if they are entirely empty
        for optional_field in ["IPC/CH area projected 2nd", "Period of 2nd projected IPC analysis"]:
            if optional_field in df_filtered.columns and df_filtered[optional_field].isna().all():
                df_filtered = df_filtered.drop(columns=optional_field)

        df_filtered = df_filtered.round(1)

        # Apply filtering only if any IPC area field is available and not all null
        filter_applied = False
        for col in [ "IPC/CH area projected 2nd", "IPC/CH area projected", "IPC/CH area current"]:
            if col in df_filtered.columns and df_filtered[col].notna().any():
                df_filtered = df_filtered[df_filtered[col] >= 3]
                filter_applied = True
                break  # Stop after applying the first valid filter


        # Drop unnecessary fields
        df_filtered = df_filtered.drop(columns=["adm0_name", "adm0_ISO3"], errors="ignore")

        # Reorder adm2 fields before adm1
        desired_order = [
            "adm2_name", "adm2_pcode",
            "adm1_name"
        ] + [col for col in df_filtered.columns if col not in ["adm1_name", "adm1_pcode", "adm2_name", "adm2_pcode"]]
        df_filtered = df_filtered[desired_order]

        df_top10 = df_filtered.sort_values(by="Cropland exposed (Km2)", ascending=False).head(10)

        return {
            "title": base_title,
            "df": df_top10
        }

    except FileNotFoundError:
        msg = f"{base_title} — SKIPPED: {adm0_iso3} File '{file_name}' not found"
        print(msg)
    except ValueError:
        msg = f"{base_title} — SKIPPED: {adm0_iso3} Sheet '{country_iso3}' not found"
        print(msg)
    except Exception as e:
        msg = f"{base_title} — SKIPPED: {adm0_iso3} Error processing sheet '{country_iso3}': {e}"
        failed_icp_floods.append(msg)
        print(msg)

    return {
        "title": msg,
        "df": pd.DataFrame()
    }

def residency_sample_size_summary(df):
    """
    Returns a clean 2-column DataFrame showing the sample size by residency group.
    """
    # Load official residency labels from metadata
    residency_labels = get_indicator_info('hh_residencetype').get("codes", {})

    # Ensure hh_residencetype is numeric
    df = df.copy()
    df["hh_residencetype"] = pd.to_numeric(df["hh_residencetype"], errors="coerce").astype("Int64")

    rows = []
    total = 0

    for code_str, label in residency_labels.items():
        if label in ["Don't know", "Refused"]:
            continue
        try:
            code = int(code_str)
        except ValueError:
            continue

        count = df[df["hh_residencetype"] == code].shape[0]
        rows.append({
            "Residency group": label,
            "Observations": count
        })
        total += count

    # Print summary in console
    # print("Sample size by residency status – Total:", total)
    # for row in rows:
    #     print(f"{row['Residency group']}: {row['Sample size']}")

    return {
        "title": "Observations by residency status. The next three tables and charts will be based on the residency status of the household. In some cases, specific groups may be underrepresented; therefore, the analysis outcomes should be interpreted with caution.",
        "metadata": f"Total sample size: {total}",
        "df": pd.DataFrame(rows)
    }

def query_shocks_trend_adm0_with_averages_and_deviation(adm0_iso3):
    from arcgis.features import FeatureLayer
    import pandas as pd

    layer_url = "https://services5.arcgis.com/sjP4Ugu5s0dZWLjd/arcgis/rest/services/diem_trend_adm0/FeatureServer/66"
    layer = FeatureLayer(layer_url)
    where_clause = f"adm0_iso3 = '{adm0_iso3}' AND indicator LIKE 'shock_%'"
    features = layer.query(where=where_clause, out_fields="*", return_geometry=False)
    df = features.sdf if features and features.features else pd.DataFrame()

    if df.empty:
        print(f"No data found for {adm0_iso3}")
        return df, pd.DataFrame(), pd.DataFrame()

    # Clean and filter
    df = df.copy()
    df["round"] = pd.to_numeric(df["round"], errors="coerce")
    df["value"] = pd.to_numeric(df["value"], errors="coerce")
    df = df.dropna(subset=["round", "value", "indicator"])
    df = df[df["indicator"] != "shock_anyshock"]

    # Identify indicators that exist in the latest round only
    latest_round_global = df["round"].max()
    indicators_in_latest_round = df[df["round"] == latest_round_global]["indicator"].unique()

    # Averages over last 6 rounds (full df, filtered after)
    last6_df = (
        df.sort_values("round", ascending=False)
          .groupby("indicator")
          .head(6)
    )
    averages = (
        last6_df.groupby("indicator")["value"]
        .mean().round(2)
        .reset_index()
        .rename(columns={"value": "avg_last6rounds"})
    )

    # Latest value per indicator
    latest_df = (
        df.sort_values("round", ascending=False)
          .groupby("indicator")
          .first()
          .reset_index()[["indicator", "value", "round"]]
          .rename(columns={"value": "latest_value", "round": "latest_round"})
    )

    # Merge and compute metrics
    merged = averages.merge(latest_df, on="indicator", how="inner")
    merged["deviation"] = (merged["latest_value"] - merged["avg_last6rounds"]).round(2)
    merged["percent_change"] = (
        ((merged["latest_value"] - merged["avg_last6rounds"]) / merged["avg_last6rounds"]) * 100
    ).round(1)

    # === Identify remarkable shocks with reasons ===
    remarkable = []

    # 1. Top 5 by average
    top_avg = averages.sort_values("avg_last6rounds", ascending=False).head(5)
    for _, row in top_avg.iterrows():
        avg = row["avg_last6rounds"]
        remarkable.append({
            "indicator": row["indicator"],
            "reason": f"Among the 5 most common shocks (avg. affected HH over last 6 rounds on Admin 0 level = {avg:.1f}%)"
        })


    # 2. Deviation ≥ 10 pp
    for _, row in merged.iterrows():
        if row["deviation"] >= 10:
            remarkable.append({
                "indicator": row["indicator"],
                "reason": f"Significant increase at adm0 level in last round: +{row['deviation']} pp compared to average"
            })

    # 3. % increase ≥ 30%
    for _, row in merged.iterrows():
        if row["percent_change"] >= 30:
            remarkable.append({
                "indicator": row["indicator"],
                "reason": f"Significant increase at adm0 level: +{row['percent_change']}% compared to average"
            })

    # Group by indicator and summarize reasons
    from collections import defaultdict
    remarkable_dict = defaultdict(list)
    for item in remarkable:
        remarkable_dict[item["indicator"]].append(item["reason"])

    # Create final output list with label
    remarkable_shocks = []
    for indicator, reasons in remarkable_dict.items():
        label = shock_label_map.get(indicator, indicator.replace("shock_", "").replace("_", " ").capitalize())
        remarkable_shocks.append({
            "indicator": indicator,
            "label": label,
            "reasons": reasons
        })

    # Filter all outputs to indicators available in the latest round
    averages = averages[averages["indicator"].isin(indicators_in_latest_round)]
    latest_df = latest_df[latest_df["indicator"].isin(indicators_in_latest_round)]
    merged = merged[merged["indicator"].isin(indicators_in_latest_round)]

    # Sort
    averages_sorted = averages.sort_values("avg_last6rounds", ascending=False).reset_index(drop=True)
    deviations_sorted = merged.sort_values("deviation", ascending=False).reset_index(drop=True)

    return df, averages_sorted, deviations_sorted, remarkable_shocks

def generate_remarkable_shocks_maps(remarkable_shocks, adm0_iso3, round_num):
    import matplotlib.pyplot as plt
    import matplotlib.colors as mcolors
    import geopandas as gpd
    import numpy as np
    from arcgis.features import FeatureLayer
    from arcgis.geometry import Geometry
    import pandas as pd

    # Create output folder
    out_dir = os.path.join("outputs_for_erps", "shocks_maps")
    os.makedirs(out_dir, exist_ok=True)

    # Load FeatureLayer and query
    url = "https://services5.arcgis.com/sjP4Ugu5s0dZWLjd/arcgis/rest/services/diem_adm_repr_1_mview/FeatureServer/29"
    layer = FeatureLayer(url)
    where_clause = f"adm0_iso3 = '{adm0_iso3}' AND round = {round_num}"
    features = layer.query(where=where_clause, out_fields="*", return_geometry=True)

    if not features or not features.features:
        print(f"[WARN] No data available for {adm0_iso3} R{round_num}")
        return

    # Convert to GeoDataFrame manually
    records = []
    for feat in features.features:
        attr = feat.attributes
        geom = feat.geometry
        if geom:
            records.append({**attr, "geometry": Geometry(geom).as_shapely})

    if not records:
        print(f"[SKIP] No geometries found for {adm0_iso3} R{round_num}")
        return

    gdf = gpd.GeoDataFrame(records, crs="EPSG:4326")

    # Load ADM0 base boundary (as basemap, not in legend)
    adm0_url = "https://services5.arcgis.com/sjP4Ugu5s0dZWLjd/arcgis/rest/services/Administrative_Boundaries_Reference_(view_layer)/FeatureServer/2"
    adm0_layer = FeatureLayer(adm0_url)
    adm0_features = adm0_layer.query(where=f"adm0_iso3 = '{adm0_iso3}'", out_fields="*", return_geometry=True)

    adm0_records = []
    for feat in adm0_features.features:
        attr = feat.attributes
        geom = feat.geometry
        if geom:
            adm0_records.append({**attr, "geometry": Geometry(geom).as_shapely})

    adm0_gdf = gpd.GeoDataFrame(adm0_records, crs="EPSG:4326") if adm0_records else None


    # Plot each remarkable shock
    for shock in remarkable_shocks:
        field = shock["indicator"] + "_1"  # e.g., "shock_flood"
        label = shock["label"]       # e.g., "Flood"

        if field not in gdf.columns:
            print(f"[SKIP] Field '{field}' not in layer")
            continue

        valid_data = gdf[[field]].dropna()
        if valid_data.empty:
            print(f"[SKIP] No valid data for '{field}'")
            continue

        values = valid_data[field].dropna().values

        if len(values) < 2 or np.nanmin(values) == np.nanmax(values):
            print(f"[SKIP] Cannot plot '{field}': not enough variation (min = max = {np.nanmin(values)})")
            continue

        quintiles = np.nanpercentile(values, [0, 20, 40, 60, 80, 100])
        if not np.all(np.isfinite(quintiles)) or np.unique(quintiles).size < 2:
            print(f"[SKIP] Invalid quintile boundaries for '{field}'")
            continue

        cmap = plt.get_cmap("YlOrRd")
        norm = mcolors.BoundaryNorm(quintiles, cmap.N, clip=True)


        fig, ax = plt.subplots(1, 1, figsize=(10, 8))
        # Plot ADM0 boundary first (gray background)
        if adm0_gdf is not None:
            adm0_gdf.boundary.plot(ax=ax, linewidth=0.8, edgecolor='lightgray')

        # Plot shock data on top
        gdf.plot(column=field, cmap=cmap, linewidth=0.3, ax=ax, edgecolor='gray', norm=norm, legend=True)

        ax.set_title(f"Percentage of population reporting {label} – {adm0_iso3} R{round_num}", fontsize=14)
        ax.axis("off")

        filename = f"{adm0_iso3}_{round_num}_{field}.png"
        out_path = os.path.join(out_dir, filename)
        plt.tight_layout()
        plt.savefig(out_path, dpi=300)
        plt.close()




# === LOAD CSV ===



if dev_mode:
    csv_path = r"C:\git\crossview_processing\DIEM_micro20250703_CODR9.csv"
else:
    csv_path = r"C:\git\crossview_processing\DIEM_micro20250729.csv"

columns_to_use = [
    "adm0_iso3", "adm0_name", "round", "weight_final", "p_mod", "p_sev",
    "hh_residencetype", "hh_agricactivity", "assistance_quality", "need"
]
# Load only relevant fields from microdata, to improve performance
# Dynamically collect all columns starting with 'need_' and 'need_received_'
sample_cols = pd.read_csv(csv_path, nrows=0).columns.tolist()
columns_to_use += [col for col in sample_cols if col.startswith("need_") or col.startswith("need_received_")]
print("Loading CSV")
df_all = pd.read_csv(csv_path, usecols=columns_to_use)
print("CSV Loaded")
#to load all csv instead (currently it takes 4 seconds more)
#df_all = pd.read_csv(csv_path)
failed_icp_floods = []
# === Main loop ===
for survey in survey_list:
    adm0_iso3 = survey["adm0_iso3"]
    adm0_name = survey["adm0_name"]
    round_num = survey["round_num"]
    diem_survey_coverage = survey["diem_survey_coverage"]
    diem_target_pop = survey["diem_target_pop"]
    coll_end_date = survey["coll_end_date"]
    print("Processing %s R%s" % (adm0_iso3, round_num))
# for survey in survey_list:
#     adm0_iso3 = survey["adm0_iso3"]
#     round_num = survey["round_num"]
    df = df_all[(df_all["adm0_iso3"] == adm0_iso3) & (df_all["round"] == round_num)]
    # print(f"Fetched {len(df)} records for {adm0_iso3} round {round_num}.")



    # === MAIN EXECUTION ===
    result_dfs = []

    # #FIES by agricultural dependancy simplified
    print('FIES by agricultural dependancy simplified')
    result_dfs.append(fies_by_simplified_agriculture(df))
    # #FIES by agricultural dependancy
    print('FIES by agricultural dependancy')
    result_dfs.append(fies_by_indicator("fies_hhtype", df))
    if "hh_residencetype" in df.columns and df["hh_residencetype"].dropna().any():
        result_dfs.append(residency_sample_size_summary(df))
        ##FIES by residency status
        print('FIES by residency status')
        result_dfs.append(fies_by_indicator("fies_resid", df))
        # #agricultural dependancy simplified by residency status
        print('agricultural dependancy simplified by residency status')
        result_dfs.append(simplified_dependency_by_residency(df))
        # #agricultural dependancy by residency status
        print('agricultural dependancy by residency status')
        result_dfs.append(agricultural_dependency(df))
    else:
        msg = "This survey did not contain a question on residency status (hh_residencetype), so residency-based analysis was skipped."
        #print(msg)
        result_dfs.append({
            "title": "Residency-based analysis",
            "metadata": None,
            "df": pd.DataFrame({"Message": [msg]})
        })
    # # Needs summary grouped, using previous round, and custom universe
    if adm0_iso3 in ['PSE']:
        print ("WARNING: Skipping Needs Analysis for PSE since a non standard questionnaire was used")
    else:
        print('Needs summary grouped, using previous round')
        needs_res = needs_summary_grouped(df_all, adm0_iso3, round_num,use_grouping=True, use_previous_round=True,universe_filter=[0, 1, 888])
        # # Assistance summary (grouped)
        print('Assistance summary (grouped)')
        assistance_res = assistance_summary(    df, use_grouping=True,universe_filter=[0, 1, 888],round_num=round_num,adm0_iso3=adm0_iso3)
        result_dfs.append(needs_res)
        result_dfs.append(assistance_res)
        # Append comparison
        print('Comparison previous needs VS assistance')
        result_dfs.append(compare_needs_vs_assistance(needs_res, assistance_res))
        # Grouped by received assistance type
        if "assistance_quality" in df.columns and df["assistance_quality"].dropna().any():
            print('Assistance quality')
            result_dfs.append(assistance_quality_summary(df, group_by="need_received"))
        else:
            msg = "This survey did not contain a question on quality of assistance, so analysis on assistance satisfaction was skipped."
            # print(msg)
            result_dfs.append({
                "title": "Satisfaction with assistance by group of assistance received",
                "metadata": None,
                "df": pd.DataFrame({"Message": [msg]})
            })


    try:
        print('IPC and flood exposure')
        result_dfs.append(extract_top10_by_cropland(adm0_iso3))
    except:
        print ('FAILED IPC TABLE INSERTION')
        failed_icp_floods.append(adm0_iso3)

    print('Shocks trends and deviations')
    df_trend, df_avg6, df_dev, remarkable_shocks = query_shocks_trend_adm0_with_averages_and_deviation(adm0_iso3)

    if not df_avg6.empty:
        df_fixed = df_avg6.rename(columns={"indicator": "Shock", "avg_last6rounds": "Weighted % of HH"}).sort_values(
            "Weighted % of HH", ascending=False)
        df_fixed["Weighted % of HH"] = df_fixed["Weighted % of HH"].astype(float)

        # Rename shocks
        df_fixed["Shock"] = df_fixed["Shock"].replace(shock_label_map)

        # Rounds metadata logic
        round_min = df_trend['round'].min()
        round_max = df_trend['round'].max()
        round_min_date = df_trend['coll_end_date'].min()
        round_max_date = df_trend['coll_end_date'].max()
        round_max_date_str = round_max_date.strftime("%b %Y")
        round_min_date_str = round_min_date.strftime("%b %Y")
        round_diff = round_max - round_min + 1  # inclusive
        if round_diff > 6:
            title_rounds = 6
            metadata_start = round_max - 5
        else:
            title_rounds = round_diff
            metadata_start = round_min

        result_dfs.append({
            "title": f"Most frequent shocks – average over last {title_rounds} rounds",
            "metadata": f"{adm0_name} – DIEM data, rounds {metadata_start} to {round_max} ({round_min_date_str} to {round_max_date_str}) ",
            "df": df_fixed
        })

    if not df_dev.empty:
        df_dev["deviation"] = df_dev["deviation"].astype(float)
        df_dev["percent_change"] = df_dev["percent_change"].astype(float)

        df_deviation = df_dev[["indicator", "deviation"]].rename(
            columns={"indicator": "Shock", "deviation": "Deviation (pp)"}
        ).sort_values("Deviation (pp)", ascending=False)

        df_percent = df_dev[["indicator", "percent_change"]].rename(
            columns={"indicator": "Shock", "percent_change": "Percent change (%)"}
        ).sort_values("Percent change (%)", ascending=False)

        # Apply renaming here too
        for df_out in [df_deviation, df_percent]:
            df_out["Shock"] = df_out["Shock"].replace(shock_label_map)

        result_dfs.append({
            "title": "Shocks: deviation from average",
            "metadata": f"{adm0_name} – DIEM data, last round: {df_dev['latest_round'].max()}",
            "df": df_deviation
        })

        result_dfs.append({
            "title": "Shocks: percent change from average",
            "metadata": f"{adm0_name} – DIEM data, last round: {df_dev['latest_round'].max()}",
            "df": df_percent
        })

    generate_remarkable_shocks_maps(remarkable_shocks, adm0_iso3, round_num)

    if export_csv:
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.chart import BarChart, Reference
        from openpyxl.styles import Font, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "DIEM Surveys analysis"
        current_row = 1

        # === Add coverage map at the top ===
        coverage_map_dir = os.path.join("outputs_for_erps", "coverage_maps")
        search_pattern = f"map_{adm0_iso3.lower()}_round{round_num}_*.png"
        matching_files = glob.glob(os.path.join(coverage_map_dir, search_pattern))
        if matching_files:
            img_path = matching_files[0]
            img = ExcelImage(img_path)

            # Open image using PIL to get original dimensions
            with PILImage.open(img_path) as pil_img:
                orig_w, orig_h = pil_img.size

            # Set a maximum width to avoid oversized images
            max_width = 900
            if orig_w > max_width:
                scale_factor = max_width / orig_w
                img.width = int(orig_w * scale_factor)
                img.height = int(orig_h * scale_factor)
            else:
                img.width = orig_w
                img.height = orig_h

            ws.add_image(img, "A1")

            # Roughly estimate rows occupied (Excel row ≈ 20 px high)
            current_row += int(img.height / 20) + 5
        else:
            print(f"No coverage map found for {adm0_iso3} R{round_num}")

        # Set wider column widths for A–E
        for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J","K","L","M",'N']:
            ws.column_dimensions[col].width = 24

        label_max_len = 30   # Truncate chart labels beyond this length

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        def truncate_label(label, max_len):
            return label if len(label) <= max_len else label[:max_len - 3] + "..."

        # Add main title
        coll_end_date_raw = survey.get("coll_end_date", "")
        try:
            coll_end_date = pd.to_datetime(coll_end_date_raw).strftime("%Y-%m-%d")
        except:
            coll_end_date = coll_end_date_raw  # fallback in case parsing fails

        title_string = f"DIEM surveys analysis for ERPs – {adm0_name} – Round {round_num} – Data collection ended on {coll_end_date}"
        ws.cell(row=current_row, column=1, value=title_string)
        ws.cell(row=current_row, column=1).font = Font(size=14, bold=True)
        current_row += 2  # leave a blank row after main title

        for i, result in enumerate(result_dfs):
            if i > 0:
                current_row += 15  # space between tables
            #deals with a special case of a missing flood exposure
            if 'is empty' in result["title"].lower():
                continue #we are dealing with a country with no flood exposure data
            # Title and metadata
            if "title" in result:
                ws.cell(row=current_row, column=1, value=result["title"])
                ws.cell(row=current_row, column=1).font = Font(bold=True)
                current_row += 1
            if "metadata" in result:
                ws.cell(row=current_row, column=1, value=result["metadata"])
                current_row += 1

            #df = result["df"].reset_index()
            df = result["df"]
            if df.index.name is not None or isinstance(df.index, pd.MultiIndex):
                df = df.reset_index()

            # === Handle special cases:
            # skip chart and optionally skip full table rendering
            contains_chart = True
            if (
                    df.shape[1] == 2 and df.columns.tolist() == ['Residency group', 'Observations']
            ):
                # Write the table as normal
                table_cols = list(df.columns)
                for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
                    for col_idx, val in enumerate(row, 1):
                        cell = ws.cell(row=current_row, column=col_idx, value=val)
                        cell.border = thin_border
                    current_row += 1

                # Do NOT generate chart for this table
                contains_chart = False
                continue
            # === Legacy case: single-row skipped messages
            if df.shape == (1, 2) and (
                    "skipped" in str(df.iloc[0, 1]).lower()
            ):
                message_text = str(df.iloc[0, 1])
                ws.cell(row=current_row, column=1, value=message_text)
                current_row += 2
                contains_chart = False
                continue

            # Remove 'None' category from assistance received charts
            if "assistance received" in result["title"].lower():
                df = df[~df.iloc[:, 0].str.strip().str.lower().eq("none")]

            # Create truncated label column for chart X-axis
            chart_labels_col = df.columns[0] + "_short"
            df[chart_labels_col] = df.iloc[:, 0].apply(lambda x: truncate_label(str(x), label_max_len))

            start_row = current_row

            # Write table (full-length labels)
            table_cols = list(df.columns[:-1])  # exclude chart_labels_col

            if "shocks" in result['title'].lower():
                # Write header
                for col_idx, col_name in enumerate(table_cols, 1):
                    ws.cell(row=current_row, column=col_idx, value=col_name).border = thin_border
                current_row += 1

                # Determine the actual column names
                first_col, second_col = table_cols

                # Write data rows manually
                for _, row in df[table_cols].iterrows():
                    ws.cell(row=current_row, column=1, value=str(row[first_col])).border = thin_border
                    ws.cell(row=current_row, column=2, value=row[second_col]).border = thin_border
                    current_row += 1

            else: #using default dataframe_to_rows for all other cases
                for row_idx, row in enumerate(dataframe_to_rows(df[table_cols], index=False, header=True)):
                    for col_idx, val in enumerate(row, 1):
                        cell = ws.cell(row=current_row, column=col_idx, value=val)
                        cell.border = thin_border
                    current_row += 1

            # Write truncated labels far right in column AZ (col 52)
            chart_label_col_idx = 52
            ws.cell(row=start_row, column=chart_label_col_idx, value="Chart labels")
            for r in range(df.shape[0]):
                label_val = df.iloc[r, df.columns.get_loc(chart_labels_col)]
                ws.cell(row=start_row + 1 + r, column=chart_label_col_idx, value=label_val)

            # Create chart
            num_cols = len(table_cols)
            num_rows = df.shape[0]
            if num_cols >= 2:

                # === Custom chart for flood exposure table (cropland + population)
                # === Custom chart for flood exposure table (cropland only, sorted)
                if result["title"].startswith(
                        "List of adm2 units in IPC3+ with the highest amount of cropland exposed to floods"):
                    try:
                        crop_col_idx = table_cols.index("Cropland exposed (Km2)")
                        adm2_name_idx = table_cols.index("adm2_name")
                    except ValueError as e:
                        print(f"Missing expected column for flood chart: {e}")
                        continue

                    # Extract data from existing table
                    data_rows = []
                    for i in range(num_rows):
                        row_idx = start_row + 1 + i  # Skip header
                        adm2 = ws.cell(row=row_idx, column=adm2_name_idx + 1).value
                        crop_val = ws.cell(row=row_idx, column=crop_col_idx + 1).value
                        data_rows.append((row_idx, adm2, crop_val))

                    # Sort rows by cropland exposed descending
                    sorted_rows = sorted(data_rows, key=lambda x: (x[2] if x[2] is not None else 0), reverse=True)

                    # Prepare references to the sorted row indexes
                    sorted_row_indices = [r[0] for r in sorted_rows]

                    # Create chart data and category references using sorted order
                    categories = Reference(ws, min_col=adm2_name_idx + 1,
                                           min_row=sorted_row_indices[0],
                                           max_row=sorted_row_indices[-1])
                    crop_data = Reference(ws, min_col=crop_col_idx + 1,
                                          min_row=sorted_row_indices[0] - 1,  # include header
                                          max_row=sorted_row_indices[-1])

                    # Create chart
                    chart_crop = BarChart()
                    chart_crop.type = "col"
                    chart_crop.grouping = "clustered"
                    chart_crop.title = "Cropland exposed to floods (top 10 adm2 in IPC3+)"
                    chart_crop.x_axis.title = "Admin2 unit"
                    chart_crop.y_axis.title = "Cropland exposed (Km²)"
                    chart_crop.y_axis.majorGridlines = None
                    chart_crop.width = 18
                    chart_crop.height = 9

                    chart_crop.add_data(crop_data, titles_from_data=True)
                    chart_crop.set_categories(categories)

                    # Insert into sheet
                    chart_row = current_row + 2
                    chart_position = f"B{chart_row}"
                    ws.add_chart(chart_crop, chart_position)

                    # Advance row counter
                    current_row = chart_row + 20




                # === Default chart for all other tables
                else:
                    chart = BarChart()
                    chart.type = "col"
                    chart.title = result["title"]
                    chart.y_axis.title = "Percentage of households"
                    # Make shock charts larger
                    if "shock" in result["title"].lower():
                        chart.width = 30  # double width
                        chart.height = 9  # double height
                    else:
                        chart.width = 18
                        chart.height = 9

                    data = Reference(ws, min_col=2, min_row=start_row,
                                     max_col=num_cols, max_row=start_row + num_rows)
                    categories = Reference(ws, min_col=chart_label_col_idx,
                                           min_row=start_row + 1, max_row=start_row + num_rows)

                    chart.add_data(data, titles_from_data=True)
                    # === Apply custom colors for satisfaction chart ===
                    if "satisfaction" in result["title"].lower():
                        custom_colors = [
                            "00B050",  # Green – Yes - satisfied
                            "C00000",  # Dark red
                            "F79646",  # Orange
                            "8B4513",  # Brown
                            "800080",  # Purple
                            "9370DB"  # Light purple
                        ]

                        for i, series in enumerate(chart.series):
                            color = custom_colors[i] if i < len(custom_colors) else custom_colors[
                                i % len(custom_colors)]
                            series.graphicalProperties.solidFill = color

                    chart.set_categories(categories)

                    if result["title"] in ["Simplified agricultural dependency by residency status", "Detailed agricultural dependency by residency status"]\
                            or 'Satisfaction'.lower() in result["title"]:
                        chart.grouping = "stacked"
                        chart.overlap = 100
                        chart.y_axis.scaling.max = 100

                    chart_row = current_row + 2
                    chart_position = f"B{chart_row}"

                    ws.add_chart(chart, chart_position)
                    if contains_chart:
                        current_row = chart_row + 7
                    else:
                        current_row = chart_row - 10


        # === Add sheet with subnational shock maps ===
        ws_maps = wb.create_sheet(title="Remarkable shocks maps")
        row_idx = 1

        # Title
        ws_maps.cell(row=row_idx, column=1, value=f"Subnational maps of remarkable shocks for round {round_num}")
        ws_maps.cell(row=row_idx, column=1).font = Font(size=14, bold=True)
        row_idx += 2

        shock_map_dir = os.path.join("outputs_for_erps", "shocks_maps")

        for shock in remarkable_shocks:
            indicator = shock["indicator"]
            label = shock["label"]
            reasons = shock["reasons"]
            shock_field = indicator + "_1"
            map_filename = f"{adm0_iso3}_{round_num}_{shock_field}.png"
            map_path = os.path.join(shock_map_dir, map_filename)

            # Title of the shock
            ws_maps.cell(row=row_idx, column=1, value=f"{label}")
            ws_maps.cell(row=row_idx, column=1).font = Font(bold=True)
            row_idx += 1

            # Reasons
            for reason in reasons:
                ws_maps.cell(row=row_idx, column=1, value=f"- {reason}")
                row_idx += 1

            # Insert map if available
            if os.path.exists(map_path):
                try:
                    img = ExcelImage(map_path)
                    with PILImage.open(map_path) as pil_img:
                        orig_w, orig_h = pil_img.size

                    max_width = 600
                    if orig_w > max_width:
                        scale_factor = max_width / orig_w
                        img.width = int(orig_w * scale_factor)
                        img.height = int(orig_h * scale_factor)
                    else:
                        img.width = orig_w
                        img.height = orig_h

                    ws_maps.add_image(img, f"A{row_idx}")
                    row_idx += int(img.height / 20) + 5
                except Exception as e:
                    print(f"Error adding image for {label}: {e}")
            else:
                ws_maps.cell(row=row_idx, column=1, value="Map not available.")
                row_idx += 2

        # Add explanation of classification
        ws_maps.cell(row=row_idx, column=1, value="Classification method:")
        ws_maps.cell(row=row_idx, column=1).font = Font(bold=True)
        row_idx += 1

        classification_note = (
            "The shock indicator maps use a quintile classification scheme, which divides the observed values into five equally sized groups "
            "(0–20th, 20–40th, 40–60th, 60–80th, 80–100th percentiles). This approach highlights relative differences across subnational units "
            "within each country and round, ensuring that variation is visible even when values are concentrated within narrow ranges."
        )

        from textwrap import wrap
        for line in wrap(classification_note, width=110):
            ws_maps.cell(row=row_idx, column=1, value=line)
            row_idx += 1



        # Add sheet with methodology text for the current survey
        ws_method = wb.create_sheet(title="DIEM Survey methodology")

        # Optional: bold title
        ws_method["A1"] = f"Methodology for {adm0_name} – Round {round_num}"
        ws_method["A1"].font = Font(size=12, bold=True)

        # Write the actual methodology content starting from row 3
        methodology_text_raw = survey.get("methodology", "No methodology information available.")
        # Remove HTML tags like <br>, <b>, <u>, etc.
        methodology_cleaned = re.sub(r"<[^>]+>", "", methodology_text_raw).strip()

        # Wrap the text every 120 characters at word boundaries
        def wrap_text(text, max_len=120):
            import textwrap
            return textwrap.fill(text, width=max_len, break_long_words=False)

        wrapped_text = wrap_text(methodology_cleaned)

        # Write each line to the Excel sheet
        for i, line in enumerate(wrapped_text.split('\n'), start=3):
            ws_method.cell(row=i, column=1, value=line)

        # Define path inside the subdirectory
        output_path = os.path.join(output_dir, f"DIEM_survey_analysis_ERPs_202507_{adm0_iso3}_{round_num}.xlsx")


        wb.save(output_path)
        print(f"\nExported grouped analysis with adaptive chart layout to: {output_path}")

print('Failed IPC and flood exposure for %s' % failed_icp_floods)